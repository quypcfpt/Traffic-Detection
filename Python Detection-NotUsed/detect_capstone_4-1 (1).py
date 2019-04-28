import cv2
import os
import numpy as np
import tensorflow as tf
import sys
import matplotlib.pyplot as plt
from filestack import Client
import threading
from xlrd import open_workbook
from utils import label_map_util
from utils import visualization_utils as vis_util
import pandas as pd
from random import randint,randrange
from openpyxl import load_workbook
from operator import sub
import collections
from math import sqrt
import operator
import json
import requests
from collections import namedtuple

MODEL_NAME = 'faster_rcnn_inception_v2_data5_200k_result'
CWD_PATH = os.getcwd()
PATH_TO_CKPT = MODEL_NAME + '/frozen_inference_graph.pb'
PATH_TO_LABELS = os.path.join('data', 'labelmap2.pbtxt')
NUM_CLASSES = 5
x=[]
y=[]
z=[]
t=[]

client = Client("AwCnfiZESWaA14iEjm189z")
sys.path.append("..")

label_map = label_map_util.load_labelmap(PATH_TO_LABELS)
categories = label_map_util.convert_label_map_to_categories(label_map, max_num_classes=NUM_CLASSES, use_display_name=True)
category_index = label_map_util.create_category_index(categories)


detection_graph = tf.Graph()
with detection_graph.as_default():
    od_graph_def = tf.GraphDef()
    with tf.gfile.GFile(PATH_TO_CKPT, 'rb') as fid:
        serialized_graph = fid.read()
        od_graph_def.ParseFromString(serialized_graph)
        tf.import_graph_def(od_graph_def, name='')
    sess = tf.Session(graph=detection_graph)
    
image_tensor = detection_graph.get_tensor_by_name('image_tensor:0')
detection_boxes = detection_graph.get_tensor_by_name('detection_boxes:0')
detection_scores = detection_graph.get_tensor_by_name('detection_scores:0')
detection_classes = detection_graph.get_tensor_by_name('detection_classes:0')
num_detections = detection_graph.get_tensor_by_name('num_detections:0')

## Get TOp 5 camera from server and get resource
video_path = {}
headers = {'Content-type': 'application/json'}
url='http://localhost:8080/api/camera'
params={'page':1,'size':3}
r=requests.get(url, params=params, headers=headers)

jsonValueResult=json.loads(r.text)
resultCamera = jsonValueResult["data"]["cameraList"]
print(jsonValueResult["data"]["cameraList"])
for i in range(0,2):
    video_path.update({resultCamera[i]["id"]:resultCamera[i]["resource"]})
##===========================
    
#video_path = [CWD_PATH + "/video/street_720.MP4",CWD_PATH + "/video/IMG_0043.MOV"]
size_excel_path = open_workbook('FixSize.xls')
size_values = []

class ReadFileExcel(object):
        def __init__(self, motor, car, bus, truck, road):
                self.motor = motor
                self.car = car
                self.bus = bus
                self.truck = truck
                self.road = road
#read excel
valuesXYZ = []

number_normal = []
number_crowed = []
number_busy = []

wb = open_workbook('FinalDataSample.xlsx')
for s in wb.sheets():
  for row in range(2,s.nrows):
    if(s.cell(row,1).value!=''):
      x.append(s.cell(row,1).value)
  for row in range(2,s.nrows):
    if(s.cell(row,2).value!=''):
      y.append(s.cell(row,2).value)
  for row in range(2,s.nrows):
    if(s.cell(row,3).value!=''):
      z.append(s.cell(row,3).value)
  for row in range(2,s.nrows):
    if(s.cell(row,4).value!=''):
      t.append(s.cell(row,4).value)

arr1 = np.array(x)
arr2 = np.array(y)
arr3 = np.array(z)
arr4 = np.array(t)
itemArray = 0
while itemArray < len(x) :
  subvalue=[arr1[itemArray],arr2[itemArray],arr3[itemArray]]
  valuesXYZ.append(subvalue)
  itemArray+=1
                
with open_workbook('FixSize.xls') as f:
        for s in f.sheets():
                for row in range(1, s.nrows):
                        values = []
                        for col in range(s.ncols):
                                value  = (s.cell(row,col).value)
                                try : value = str(float(value))
                                except ValueError : pass
                                finally:
                                        values.append(value)
                        item =  ReadFileExcel(*values)       
                        size_values.append(item)
      
def length_of_bounding_box(ymin,ymax):
        return (ymax - ymin)/2 + ymin

def sortFirst(val): 
        return val[0]

class MyClass:
    def __init__(self, label, distance):
        self.label = label
        self.distance = distance

class DetectMutilThread (threading.Thread):
    def __init__(self, name, video_url):
        threading.Thread.__init__(self)
        self.name = name
        self.video_url = video_url
   
    def run(self):
        print("Starting " + self.name + " ")
        window_name = self.name
        cv2.namedWindow(window_name)
        video = cv2.VideoCapture(self.video_url)
        fps = int(video.get(cv2.CAP_PROP_FPS))
        count = 0
        pathList=[]
        joinPath = ', '
        while True:
            frame, image = video.read()
            path = './capture/' + self.name + "_" + str(count) + '.jpg'
            imgURLs = ''
            if (count%(10*fps) == 0) & (count != 0):
                print("**********")
                print('Processing...' + self.name + ' in frame ' + str(count))
                cv2.imwrite(path, image)
                new_filelink = client.upload(filepath=path, multipart=False)
                print("link : ", new_filelink.url)
                pathList.append(new_filelink.url)
                height, width, channels = image.shape
                image_expanded = np.expand_dims(image, axis=0)

                (boxes, scores, classes, num) = sess.run(
                    [detection_boxes, detection_scores, detection_classes, num_detections],
                    feed_dict={image_tensor: image_expanded})

                vis_util.visualize_boxes_and_labels_on_image_array(
                    image,
                    np.squeeze(boxes),
                    np.squeeze(classes).astype(np.int32),
                    np.squeeze(scores),
                    category_index,
                    use_normalized_coordinates=True,
                    line_thickness=5,
                    min_score_thresh=0.50)
                threshold = 0.5
                objectLine1 = []
                
                motorCountLine1=0
                carCountLine1=0
                truckCountLine1=0
                busCountLine1=0

                motorCountLine2=0
                carCountLine2=0
                truckCountLine2=0
                busCountLine2=0

                motorCountLine3=0
                carCountLine3=0
                truckCountLine3=0
                busCountLine3=0
                
                for index, value in enumerate(classes[0]):
                  object_dict = {}
                  box=[]
                  if scores[0, index] > threshold:
                    ymin = boxes[0,index][0]* height
                    xmin = boxes[0,index][1]* width
                    ymax = boxes[0,index][2]* height
                    xmax = boxes[0,index][3]* width
                    object_dict['type'] =(category_index.get(value)).get('name')
                    object_dict['yValue'] =length_of_bounding_box(ymin , ymax)
                    object_dict['xValue'] =length_of_bounding_box(xmin , xmax)
                    objectLine1.append(object_dict)

                
                for obj in objectLine1:
                    keys = iter(obj.values())
                    typeVehicle,yValue,xValue=next(keys),next(keys),next(keys)
                    motorLabel="motor"
                    carLabel="car"
                    truckLabel="truck"
                    busLabel="bus"
                    
                    if(yValue >=30 and yValue <=270 and xValue >= 50 and xValue <= 1200):
                                if(typeVehicle.strip() == motorLabel):
                                        motorCountLine1+=1
                                elif(typeVehicle.strip() == carLabel):
                                        carCountLine1+=1
                                elif(typeVehicle.strip() == truckLabel):
                                        truckCountLine1+=1
                                else:
                                        busCountLine1+=1
                                        
                    if(yValue >=290 and yValue <=470 and xValue >= 50 and xValue <= 1270):
                                if(typeVehicle.strip() == motorLabel):
                                        motorCountLine2+=1
                                elif(typeVehicle.strip() == carLabel):
                                        carCountLine2+=1
                                elif(typeVehicle.strip() == truckLabel):
                                        truckCountLine2+=1
                                else:
                                        busCountLine2+=1
                                        
                    if(yValue >=490 and yValue <=680 and xValue >= 50 and xValue <= 1300):
                                if(typeVehicle.strip() == motorLabel):
                                        motorCountLine3+=1
                                elif(typeVehicle.strip() == carLabel):
                                        carCountLine3+=1
                                elif(typeVehicle.strip() == truckLabel):
                                        truckCountLine3+=1
                                else:
                                        busCountLine3+=1 


                                                 

                image[30:270, 50:1200, 2]=255
                image[290:470, 50:1270, 2]=255
                image[490:680, 50:1300, 2]=255

                for e in size_values:
                        delta = []
                        distList = []
                        x = float(e.road) - ((float(e.motor)*motorCountLine1) + (float(e.car)*carCountLine1) + (float(e.truck)*truckCountLine1) + (float(e.bus)*busCountLine1))
                        delta.append(x)
                        y = float(e.road) - ((float(e.motor)*motorCountLine2) + (float(e.car)*carCountLine2) + (float(e.truck)*truckCountLine2) + (float(e.bus)*busCountLine2))
                        delta.append(y)
                        z = float(e.road) - ((float(e.motor)*motorCountLine3) + (float(e.car)*carCountLine3) + (float(e.truck)*truckCountLine3) + (float(e.bus)*busCountLine3))
                        delta.append(z)
##                        print("X,Y,Z is appended into delta: " + str(x) + " " + str(y) + " " + str(z) + " of " + self.name + " in frame_" + str(count))
##                        print("Current Delta: " + str(delta) + " of " + self.name + " in frame_" + str(count))
                        
                        for i in range(0,len(valuesXYZ)):
                            vector = [sub(ax,az) for ax,az in zip(delta, valuesXYZ[i])]
                            distance = sqrt((vector[0])**2 + (vector[1])**2 + (vector[2])**2)
                            distList.append(distance)
##                            print("vector: " + str(vector) + " dist: " + str(distance) + " in frame_" + str(count))
##                        print("distList: " + str(distList) + " in frame_" + str(count))
                        
                        finalData =dict()
                        my_objects = []
                        countCrowed=0
                        countNormal=0
                        countBusy=0
                        s_crowed=0
                        s_busy=0
                        s_normal=0
                        avg_crowed=0
                        avg_busy=0
                        avg_normal=0
                        countAvgCrowed=0
                        countAvgBusy=0
                        countAvgNormal=0
                        for i in range(0, len(distList)):
                          p=(distList[i], arr4[i])
                          my_objects.append(p)
                        my_objects.sort(key = sortFirst)
                        for i in range(0,181):
                                if(str(my_objects[i][1]) == 'crowed'):
                                        countCrowed+=1
                                elif(str(my_objects[i][1]) == 'normal'):
                                        countNormal+=1
                                else:
                                        countBusy+=1
                        print(self.name + ' Crowed' + ' in frame ' + str(count) + ' : ' + str(countCrowed))
                        print(self.name + ' Normal' + ' in frame ' + str(count) + ' : ' + str(countNormal))
                        print(self.name + ' Busy' + ' in frame ' + str(count) + ' : ' + str(countBusy))
                     
                        if(countCrowed > countNormal and countCrowed > countBusy):
                                number_crowed.append(countCrowed)
                                print(self.name + " status after 10s: Crowed")
                        elif(countBusy > countNormal and countBusy > countCrowed):      
                                number_busy.append(countBusy)
                                print(self.name + " status after 10s: Busy")
                        elif(countNormal>countBusy and countNormal>countCrowed):  
                                number_normal.append(countNormal)
                                print(self.name + " status after 10s: Normal")
                        else:
                                if(countCrowed > countNormal and countBusy>countNormal and countCrowed == countBusy):
                                        for i in range(0, len(my_objects)):
                                               if(str(my_objects[i][1]) == 'crowed'):
                                                       s_crowed+=my_objects[i][0]
                                                       countAvgCrowed+=1
                                                       
                                               if(str(my_objects[i][1]) == 'busy'):
                                                       s_busy+=my_objects[i][0]
                                                       countAvgBusy+=1
                                        
                                        avg_busy = s_busy/countAvgBusy
                                        avg_crowed = s_crowed/countAvgCrowed
                                        if(avg_busy<avg_crowed):
                                                number_busy.append(countBusy)
                                                print(self.name + " status after 10s: Busy")
                                        else:
                                                number_crowed.append(countCrowed)
                                                print(self.name + " status after 10s: Crowed")
                                elif(countBusy > countNormal and countNormal > countCrowed and countBusy == countNormal):
                                        for i in range(0, len(my_objects)):
                                                if(str(my_objects[i][1]) == 'normal'):
                                                       s_normal+=my_objects[i][0]
                                                       countAvgNormal+=1
                                                       
                                                if(str(my_objects[i][1]) == 'busy'):
                                                       s_busy+=my_objects[i][0]
                                                       countAvgBusy+=1
                                        avg_busy = s_busy/countAvgBusy
                                        avg_normal = s_normal/countAvgNormal
                                        if(avg_busy<avg_normal):
                                                number_busy.append(countBusy)
                                                print(self.name + " status after 10s: Busy")
                                        else:
                                                number_normal.append(countNormal)
                                                print(self.name + " status after 10s: Normal")
                                else:
                                        for i in range(0, len(my_objects)):
                                                if(str(my_objects[i][1]) == 'normal'):
                                                       s_normal+=my_objects[i][0]
                                                       countAvgNormal+=1
                                                       
                                                if(str(my_objects[i][1]) == 'crowed'):
                                                       s_crowed+=my_objects[i][0]
                                                       countAvgCrowed+=1
                                        avg_crowed = s_crowed/countAvgCrowed
                                        avg_normal = s_normal/countAvgNormal
                                        if(avg_crowed<avg_normal):
                                                number_crowed.append(countCrowed)
                                                print(self.name + " status after 10s: Crowed")
                                        else:
                                                number_normal.append(countNormal)
                                                print(self.name + " status after 10s: Normal")

##                cv2.imwrite(path, image)
##                new_filelink = client.upload(filepath=path, multipart=False)
##                print("link : ", new_filelink.url)
##                pathList.append(new_filelink.url)
                        
                        
                        
            if ((count%(60*fps) == 0) & (count != 0)) :
                    
                    print("********************")
                    print("********************")
                    print(len(pathList))
                    for i in range(0, len(pathList)):
                        print("aaaaa" + str(pathList[i]))
                    
                    imgURLs = joinPath.join(pathList)
                    status = 0
                    if(len(number_crowed) > len(number_normal) and len(number_crowed) > len(number_busy)):
                            print(self.name + " result after 60s: Crowed")
                            status = 2
                    elif(len(number_busy) > len(number_crowed) and len(number_busy) > len(number_normal)):
                            print(self.name + " result after 60s: Busy")
                            status = 1
                    else:
                            print(self.name + " result after 60s: Normal")
                            status = 0
                    number_normal.clear()
                    number_busy.clear()
                    number_crowed.clear()
                    pathList.clear()
                    jsonData ={
                        "CameraId":self.name,
                        "statusId":status,
                        "imageUrl":imgURLs
                }
                    headers = {'Content-type': 'application/json'}
                    url='http://localhost:8080/api/detection'
                    params={'detectResultString':json.dumps(jsonData)}
                    r=requests.post(url, params=params, headers=headers)
                    print(r.status_code, r.reason, r.text)
            if not frame:  # error on video source or last frame finished
                break
            
            count+=1
           
            if cv2.waitKey(2) == ord('q'):
                break
     
        video.release()
        cv2.destroyWindow(window_name)
        print(window_name + " Exiting ")
        
def main():       
    for keys,values in video_path.items():
        thread = DetectMutilThread(keys, values)
        thread.start()  
        
if __name__ == '__main__':
    main()
