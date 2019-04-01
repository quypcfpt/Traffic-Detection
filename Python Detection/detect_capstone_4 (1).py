import cv2
import os
import numpy as np
import tensorflow as tf
import sys
import matplotlib.pyplot as plt
from filestack import Client
import threading
from xlrd import open_workbook
from utils import label_map_util
from utils import visualization_utils as vis_util
import pandas as pd
from random import randint,randrange
from openpyxl import load_workbook
from operator import sub
import collections
from math import sqrt
import operator

MODEL_NAME = 'faster_rcnn_inception_v2_data5_200k_result'
CWD_PATH = os.getcwd()
PATH_TO_CKPT = MODEL_NAME + '/frozen_inference_graph.pb'
PATH_TO_LABELS = os.path.join('data', 'labelmap2.pbtxt')
NUM_CLASSES = 5
x=[]
y=[]
z=[]
t=[]
file="DataSample.xlsx"
df = pd.read_excel(file, header=None)
writer = pd.ExcelWriter(file, engine='openpyxl')
client = Client("AwCnfiZESWaA14iEjm189z")
sys.path.append("..")

label_map = label_map_util.load_labelmap(PATH_TO_LABELS)
categories = label_map_util.convert_label_map_to_categories(label_map, max_num_classes=NUM_CLASSES, use_display_name=True)
category_index = label_map_util.create_category_index(categories)

detection_graph = tf.Graph()
with detection_graph.as_default():
    od_graph_def = tf.GraphDef()
    with tf.gfile.GFile(PATH_TO_CKPT, 'rb') as fid:
        serialized_graph = fid.read()
        od_graph_def.ParseFromString(serialized_graph)
        tf.import_graph_def(od_graph_def, name='')
    sess = tf.Session(graph=detection_graph)
    
image_tensor = detection_graph.get_tensor_by_name('image_tensor:0')
detection_boxes = detection_graph.get_tensor_by_name('detection_boxes:0')
detection_scores = detection_graph.get_tensor_by_name('detection_scores:0')
detection_classes = detection_graph.get_tensor_by_name('detection_classes:0')
num_detections = detection_graph.get_tensor_by_name('num_detections:0')

video_path = [CWD_PATH + "/ket4.mp4"]
##video_path = [CWD_PATH + "/video/normal2.mp4",CWD_PATH + "/video/IMG_0043.MOV"]
size_excel_path = open_workbook('FixSize.xls')
size_values = []

class ReadFileExcel(object):
        def __init__(self, motor, car, bus, truck, road):
                self.motor = motor
                self.car = car
                self.bus = bus
                self.truck = truck
                self.road = road
#read excel
valuesXYZ = []
number_absent = []
number_normal = []
number_busy = []

wb = open_workbook('DataSample.xlsx')
for s in wb.sheets():
  print('Sheet:',s.name)
  for row in range(2,s.nrows):
    if(s.cell(row,1).value!=''):
      x.append(s.cell(row,1).value)
  for row in range(2,s.nrows):
    if(s.cell(row,2).value!=''):
      y.append(s.cell(row,2).value)
  for row in range(2,s.nrows):
    if(s.cell(row,3).value!=''):
      z.append(s.cell(row,3).value)
  for row in range(2,s.nrows):
    if(s.cell(row,4).value!=''):
      t.append(s.cell(row,4).value)

arr1 = np.array(x)
arr2 = np.array(y)
arr3 = np.array(z)
arr4 = np.array(t)
itemArray = 0
while itemArray < len(x) :
  subvalue=[arr1[itemArray],arr2[itemArray],arr3[itemArray]]
  valuesXYZ.append(subvalue)
  itemArray+=1
print(valuesXYZ)
                
with open_workbook('FixSize.xls') as f:
        for s in f.sheets():
                for row in range(1, s.nrows):
                        values = []
                        for col in range(s.ncols):
                                value  = (s.cell(row,col).value)
                                try : value = str(float(value))
                                except ValueError : pass
                                finally:
                                        values.append(value)
                        item =  ReadFileExcel(*values)       
                        size_values.append(item)
      
def length_of_bounding_box(ymin,ymax):
        return (ymax - ymin)/2 + ymin

def sortFirst(val): 
        return val[0]

def getSecond(val): 
        return val[1] 
def isDescending(list):
    previous = list[0]
    for number in list:
        if number < previous:
            return False
        previous = number
    return True

class MyClass:
    def __init__(self, label, distance):
        self.label = label
        self.distance = distance

class DetectMutilThread (threading.Thread):
    def __init__(self, name, video_url):
        threading.Thread.__init__(self)
        self.name = name
        self.video_url = video_url
   
    def run(self):
        print("Starting " + self.name + " ")
        window_name = self.name
        cv2.namedWindow(window_name)
        video = cv2.VideoCapture(self.video_url)
        fps = int(video.get(cv2.CAP_PROP_FPS))
        count = 0
        #density = []
        
        
        while True:
            frame, image = video.read()
            path = './capture/' + self.name + "_" + str(count) + '.jpg'
            if count%(3*fps) == 0 :
                print('Creating...' + path)
                height, width, channels = image.shape
                image_expanded = np.expand_dims(image, axis=0)

                (boxes, scores, classes, num) = sess.run(
                    [detection_boxes, detection_scores, detection_classes, num_detections],
                    feed_dict={image_tensor: image_expanded})

                vis_util.visualize_boxes_and_labels_on_image_array(
                    image,
                    np.squeeze(boxes),
                    np.squeeze(classes).astype(np.int32),
                    np.squeeze(scores),
                    category_index,
                    use_normalized_coordinates=True,
                    line_thickness=5,
                    min_score_thresh=0.50)
                threshold = 0.5
                objectLine1 = []
                objectLine2 = []
                objectLine3 = []
                
                motorCountLine1=0
                carCountLine1=0
                truckCountLine1=0
                busCountLine1=0

                motorCountLine2=0
                carCountLine2=0
                truckCountLine2=0
                busCountLine2=0

                motorCountLine3=0
                carCountLine3=0
                truckCountLine3=0
                busCountLine3=0
                
                for index, value in enumerate(classes[0]):
                  object_dict = {}
                  box=[]
                  if scores[0, index] > threshold:
                    #print(len(boxes[0,index]))
                    ymin = boxes[0,index][0]* height
                    xmin = boxes[0,index][1]* width
                    ymax = boxes[0,index][2]* height
                    xmax = boxes[0,index][3]* width
                    object_dict['type'] =(category_index.get(value)).get('name')
                    object_dict['yValue'] =length_of_bounding_box(ymin , ymax)
                    object_dict['xValue'] =length_of_bounding_box(xmin , xmax)
                    objectLine1.append(object_dict)
                    objectLine2.append(object_dict)
                    objectLine3.append(object_dict)
                    
                
                
                for obj in objectLine1:
                    keys = iter(obj.values())
                    typeVehicle,yValue,xValue=next(keys),next(keys),next(keys)
##                    typeVehicle,yValue=next(keys),next(keys)
                    #print(typeVehicle ,"=",yValue,"=",xValue)
                    motorLabel="motor"
                    carLabel="car"
                    truckLabel="truck"
                    busLabel="bus"
                    
                    if(yValue >=90 and yValue <=260 and xValue >= 100 and xValue <= 1200):
                        if(typeVehicle.strip() == motorLabel):
                                motorCountLine1+=1
                        elif(typeVehicle.strip() == carLabel):
                                carCountLine1+=1
                        elif(typeVehicle.strip() == truckLabel):
                                truckCountLine1+=1
                        else:
                                busCountLine1+=1



                for obj2 in objectLine2:
                        keys = iter(obj2.values())
                        typeVehicle,yValue,xvalue=next(keys),next(keys),next(keys)
                        motorLabel="motor"
                        carLabel="car"
                        truckLabel="truck"
                        busLabel="bus"
                        
                    
                        if(yValue >=290 and yValue <=460 and xValue >= 90 and xValue <= 1250):
                                if(typeVehicle.strip() == motorLabel):
                                        motorCountLine2+=1
                                elif(typeVehicle.strip() == carLabel):
                                        carCountLine2+=1
                                elif(typeVehicle.strip() == truckLabel):
                                        truckCountLine2+=1
                                else:
                                        busCountLine2+=1

                
                for obj3 in objectLine3:
                        keys = iter(obj3.values())
                        typeVehicle,yValue,xvalue=next(keys),next(keys),next(keys)
                        motorLabel="motor"
                        carLabel="car"
                        truckLabel="truck"
                        busLabel="bus"
                        
                    
                        if(yValue >=490 and yValue <=660 and xValue >= 80 and xValue <= 1300):
                                if(typeVehicle.strip() == motorLabel):
                                        motorCountLine3+=1
                                elif(typeVehicle.strip() == carLabel):
                                        carCountLine3+=1
                                elif(typeVehicle.strip() == truckLabel):
                                        truckCountLine3+=1
                                else:
                                        busCountLine3+=1                        

                font = cv2.FONT_HERSHEY_SIMPLEX

                cv2.putText(
                        image,
                        'Line 1',
                        (50,70),
                        font,
                        1,
                        (0,0,255),
                        2
                        )
                cv2.putText(
                    image,
                    'Car: ' + str(carCountLine1),
                    (50,150),
                    font,
                    1,
                    (0,0,255),
                    2,
                    )
                cv2.putText(
                    image,
                    'Motor: ' + str(motorCountLine1),
                    (50,230),
                    font,
                    1,
                    (0,0,255),
                    2,
                    )   
                cv2.putText(
                    image,
                    'Bus: ' + str(busCountLine1),
                    (50,310),
                    font,
                    1,
                    (0,0,255),
                    2,
                    )

                cv2.putText(
                    image,
                    'Truck: ' + str(truckCountLine1),
                    (50,390),
                    font,
                    1,
                    (0,0,255),
                    2,
                    )

                cv2.putText(
                        image,
                        'Line 2',
                        (250,70),
                        font,
                        1,
                        (0,0,255),
                        2
                        )
                cv2.putText(
                    image,
                    'Car: ' + str(carCountLine2),
                    (250,150),
                    font,
                    1,
                    (0,0,255),
                    2,
                    )
                cv2.putText(
                    image,
                    'Motor: ' + str(motorCountLine2),
                    (250,230),
                    font,
                    1,
                    (0,0,255),
                    2,
                    )   
                cv2.putText(
                    image,
                    'Bus: ' + str(busCountLine2),
                    (250,310),
                    font,
                    1,
                    (0,0,255),
                    2,
                    )

                cv2.putText(
                    image,
                    'Truck: ' + str(truckCountLine2),
                    (250,390),
                    font,
                    1,
                    (0,0,255),
                    2,
                    )

                cv2.putText(
                        image,
                        'Line 3',
                        (500,70),
                        font,
                        1,
                        (0,0,255),
                        2
                        )
                cv2.putText(
                    image,
                    'Car: ' + str(carCountLine3),
                    (500,150),
                    font,
                    1,
                    (0,0,255),
                    2,
                    )
                cv2.putText(
                    image,
                    'Motor: ' + str(motorCountLine3),
                    (500,230),
                    font,
                    1,
                    (0,0,255),
                    2,
                    )   
                cv2.putText(
                    image,
                    'Bus: ' + str(busCountLine3),
                    (500,310),
                    font,
                    1,
                    (0,0,255),
                    2,
                    )

                cv2.putText(
                    image,
                    'Truck: ' + str(truckCountLine3),
                    (500,390),
                    font,
                    1,
                    (0,0,255),
                    2,
                    )
##                cv2.line(image,(0,511),(width,511),(255,0,0),5)
                image[90:260, 100:1200, 2]=255
                image[290:460, 90:1250, 2]=255
                image[490:660, 80:1300, 2]=255

                #plt.figure( figsize = (10,10))
##                def avg_dict(d1,d2):
##                        for key,value in d1.items():
##                                d1[value] = (key + d2.get(0,value)) / 2
##                                return d1
         
                for e in size_values:
                        delta = []
                        distList = []
                        x = float(e.road) - ((float(e.motor)*motorCountLine1) + (float(e.car)*carCountLine1) + (float(e.car)*truckCountLine1) + (float(e.car)*busCountLine1))
                        delta.append(x)
                        y = float(e.road) - ((float(e.motor)*motorCountLine2) + (float(e.car)*carCountLine2) + (float(e.car)*truckCountLine2) + (float(e.car)*busCountLine2))
                        delta.append(y)
                        z = float(e.road) - ((float(e.motor)*motorCountLine3) + (float(e.car)*carCountLine3) + (float(e.car)*truckCountLine3) + (float(e.car)*busCountLine3))
                        delta.append(z)
                        print("X,Y,Z is appended into delta: " + str(x) + " " + str(y) + " " + str(z) + " of " + self.name + " in frame_" + str(count))
                        print("Current Delta: " + str(delta) + " of " + self.name + " in frame_" + str(count))
                        
                        for i in range(0,len(valuesXYZ)):
                            vector = [sub(ax,az) for ax,az in zip(delta, valuesXYZ[i])]
                            distance = sqrt((vector[0])**2 + (vector[1])**2 + (vector[2])**2)
                            distList.append(distance)
                            print("vector: " + str(vector) + " dist: " + str(distance) + " in frame_" + str(count))
                        print("distList: " + str(distList) + " in frame_" + str(count))
                        
                        finalData =dict()
                        my_objects = []
                        countNormal=0
                        countAbsent=0
                        countBusy=0
                        countAvg=0

                        print("distList: " + str(distList))
                        
                        for i in range(0, len(distList)):
                        ##  p = MyClass(distList[i], arr4[i])
                          p=(distList[i], arr4[i])
                          my_objects.append(p)
                        print(type(my_objects))
                        for i in range(0,50):
                                my_objects.sort(key = sortFirst)
                                print(str(my_objects[i][1]))
                                if(str(my_objects[i][1]) == 'normal'):
                                        countNormal+=1
                                        
                                elif(str(my_objects[i][1]) == 'busy'):
                                        countBusy+=1
                                      
                                else:
                                        countAbsent+=1
                                       

                        print('normal :',countNormal)
                        
                        print('absent :',countAbsent)
                        
                        print('busy: ', countBusy)
                       
                        if(countNormal > countAbsent and countNormal > countBusy):
                          print('Normal: ' + " in frame_" + str(count))
                          number_normal.append(countNormal)
                         
                        elif(countBusy > countAbsent and countBusy > countNormal):
                          print('Busy: ' + " in frame_" + str(count))
                          number_busy.append(countBusy)
                        else:
                          print('Absent: ' + " in frame_" + str(count))
                          number_absent.append(countAbsent)
                         
                        
                        
##                        array_dic = []
                      
##                        print("X,Y,Z is appended into delta: " + str(x) + " " + str(y) + " " + str(z) + " of " + self.name + " in frame_" + str(count))
##                        print("Current Delta: " + str(delta) + " of " + self.name + " in frame_" + str(count))

##                        if(countNormal==countBusy):
##                                for keys,values in od.items():
##                                        if values == 'normal':
##                                                array_dic.append(keys)
##                                
##                                                print("array_dic: ", array_dic)
                                
                cv2.imwrite(path, image)
              
                

            if ((count%(30*fps) == 0) & (count != 0)) :
                    print('SAU 30S:')
                    print(len(number_normal))
                    print(len(number_busy))
                    if(len(number_normal) > len(number_absent) and len(number_normal) > len(number_busy)):
                            print("Normal")
                    elif(len(number_busy) > len(number_normal) and len(number_busy) > len(number_absent)):
                            print("Busy")
                    else:
                            print("Absent")
                    number_busy.clear()
                    number_normal.clear()
                    print(len(number_normal))
                    print(len(number_busy))
##                    currentTime = str(datetime.datetime.now().time())
##                    jsonData ={
##                        "CameraId":1,
##                        "statusId":status,
##                        "imageUrl":'AAA'
##                }
##                    
##                    headers = {'Content-type': 'application/json'}
##                    url='http://localhost:8080/api/detection'
##                    params={'detectResultString':json.dumps(jsonData)}
##                    r=requests.post(url, params=params, headers=headers)
##                    print(r.status_code, r.reason, r.text)
                
            if not frame:  # error on video source or last frame finished
                break
            
            count+=1
           
            if cv2.waitKey(2) == ord('q'):
                break
     
        video.release()
        cv2.destroyWindow(window_name)
##        save_Excel()
        print(window_name + " Exiting ")

def save_Excel():
    df2 = pd.DataFrame({'Data': dataX})
    df1 = pd.DataFrame({'Data': dataY})
    df3 = pd.DataFrame({'Data': dataZ})
   
    book=load_workbook(file)
    print('**********************************************************')
    print(dataX)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df2.to_excel(writer, sheet_name='Sheet1', header=None, index=False,
             startcol=1,startrow=2)
    df1.to_excel(writer, sheet_name='Sheet1', header=None, index=False,
             startcol=2,startrow=2)
    df3.to_excel(writer, sheet_name='Sheet1', header=None, index=False,
             startcol=3,startrow=2)
   
    writer.save()
        
def main():       
    for item in video_path:
        thread = DetectMutilThread(os.path.basename(item), item)
        thread.start()  
        
if __name__ == '__main__':
    main()
